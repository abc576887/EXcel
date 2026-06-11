import streamlit as st
import pandas as pd
import plotly.express as px
import io
from openpyxl.styles import PatternFill # Excel ထဲတွင် အရောင်ခြယ်ရန်

# 1. Page Configuration
st.set_page_config(page_title="Color-Coded Audit Engine", page_icon="🎨", layout="wide")

st.title("🎨 Color-Coded Advanced Audit Automation Engine")
st.markdown("အမှားများကို **ဇယားပေါ်တွင် အရောင်ဖြင့် ချက်ချင်းပြသပေးပြီး** တည်းဖြတ်ကာ Excel ထုတ်လျှင်လည်း အရောင်ပါရှိမည့် စနစ် ဖြစ်ပါသည်။")
st.markdown("---")

# 2. Sidebar Upload
st.sidebar.header("📁 Data Upload Center")
uploaded_file = st.sidebar.file_uploader("စစ်ဆေးမည့် Excel Ledger ဖိုင်ကို တင်ပါ", type=["xlsx"])

if uploaded_file is not None:
    try:
        excel_file = pd.ExcelFile(uploaded_file)
        sheet_names = excel_file.sheet_names
        
        # --- POWER QUERY STYLE DYNAMIC MAPPING ---
        sheet_mappings = {}
        df_list = []
        
        for sheet in sheet_names:
            preview_df = pd.read_excel(uploaded_file, sheet_name=sheet, nrows=2)
            cols = list(preview_df.columns)
            
            with st.sidebar.expander(f"📄 Sheet: {sheet} Columns"):
                m_vch = st.selectbox("Voucher/ID", cols, key=f"vch_{sheet}")
                m_acc = st.selectbox("Account Name", cols, key=f"acc_{sheet}")
                m_deb = st.selectbox("Debit", cols, key=f"deb_{sheet}")
                m_cre = st.selectbox("Credit", cols, key=f"cre_{sheet}")
                
                sheet_mappings[sheet] = {'vch': m_vch, 'acc': m_acc, 'deb': m_deb, 'cre': m_cre}
        
        for sheet in sheet_names:
            raw_df = pd.read_excel(uploaded_file, sheet_name=sheet)
            if not raw_df.empty:
                maps = sheet_mappings[sheet]
                processed_df = pd.DataFrame()
                processed_df['Voucher_No'] = raw_df[maps['vch']].astype(str).str.strip()
                processed_df['Account_Name'] = raw_df[maps['acc']].astype(str).str.strip()
                processed_df['Debit'] = pd.to_numeric(raw_df[maps['deb']], errors='coerce').fillna(0)
                processed_df['Credit'] = pd.to_numeric(raw_df[maps['cre']], errors='coerce').fillna(0)
                processed_df['Sheet_Source'] = sheet
                
                for col in raw_df.columns:
                    if col not in maps.values():
                        processed_df[f"Orig_{col}"] = raw_df[col]
                df_list.append(processed_df)
                
        combined_df = pd.concat(df_list, ignore_index=True)
        
        st.markdown("---")
        
        # --- STEP 2: LIVE DATA REVIEW & EDITING ---
        st.subheader("📝 Step 1: Live Review & Edit Ledger Data")
        st.markdown("💡 အောက်ပါ ဇယားတွင် ဒေတာများကို တိုက်ရိုက်ကလစ်နှိပ်၍ ပြင်ဆင်နိုင်ပါသည်။")
        
        edited_df = st.data_editor(combined_df, num_rows="dynamic", use_container_width=True, key="ledger_editor")
        
        st.markdown("---")
        
        # --- STEP 3: COLOR HIGHLIGHTING LOGIC ---
        st.subheader("🔍 Step 2: Automated Audit Rules & Color Coding")
        st.info("🚨 **အရောင်သတ်မှတ်ချက်:** [ အနီရောင် = Debit/Credit လွဲမှားမှု ] | [ အဝါရောင် = Voucher နံပါတ် ထပ်နေမှု ] | [ မီးခိုးရောင် = Account နာမည်မပါမှု ]")
        
        # စာကြောင်းအလိုက် အရောင်ခြယ်မည့် Python Function (Streamlit UI အတွက်)
        def highlight_audit_errors(row):
            styles = [''] * len(row)
            
            # Rule 1: Unbalanced (Debit == Credit and not 0) သို့မဟုတ် (နှစ်ခုလုံး 0 ဖြစ်နေလျှင်) -> အနီရောင်
            if ((row['Debit'] == row['Credit']) & (row['Debit'] != 0)) or ((row['Debit'] == 0) & (row['Credit'] == 0)):
                return ['background-color: #ffcccc; color: black'] * len(row)
                
            # Rule 2: Account Name ကျန်ခဲ့လျှင် -> မီးခိုးရောင်
            if pd.isna(row['Account_Name']) or row['Account_Name'] in ['nan', '']:
                return ['background-color: #f0f0f0; color: black'] * len(row)
                
            return styles

        # စတိုင်ကို Apply လုပ်ပြီး UI ပေါ်တွင် ပြသခြင်း
        # Voucher ထပ်တာကိုတော့ ဇယားတစ်ခုလုံးအတိုင်း စစ်မှရလို့ သီးသန့်ကြည့်ရှုနိုင်အောင် Tabs နဲ့ပါ ပြထားပေးပါတယ်
        duplicated_vouchers = edited_df[edited_df.duplicated(subset=['Voucher_No'], keep=False) & (edited_df['Voucher_No'] != 'nan')]['Voucher_No'].tolist()
        
        def highlight_cells(df):
            # ပုံသေ DataFrame တစ်ခုဆောက်
            style_df = pd.DataFrame('', index=df.index, columns=df.columns)
            for idx, row in df.iterrows():
                # Unbalanced Check -> Red
                if ((row['Debit'] == row['Credit']) & (row['Debit'] != 0)) or ((row['Debit'] == 0) & (row['Credit'] == 0)):
                    style_df.loc[idx] = 'background-color: #ffcccc; color: #800000; font-weight: bold;'
                # Missing Account -> Gray
                elif pd.isna(row['Account_Name']) or row['Account_Name'] in ['nan', '']:
                    style_df.loc[idx] = 'background-color: #e0e0e0; color: #555555;'
                # Duplicate Voucher -> Yellow
                elif row['Voucher_No'] in duplicated_vouchers:
                    style_df.loc[idx] = 'background-color: #fff2cc; color: #b2a100;'
            return style_df

        styled_output = edited_df.style.apply(highlight_cells, axis=None)
        st.dataframe(styled_output, use_container_width=True)
        
        # --- EXPORT TO EXCEL WITH COLORS ---
        st.markdown("---")
        st.subheader("📥 Step 3: Export Color-Coded Excel Report")
        st.markdown("ဒေါင်းလုဒ်လုပ်မည့် Excel ဖိုင်ထဲတွင်လည်း **အမှားများကို အရောင်ခြယ်ပြီးသား** တစ်ခါတည်း ပါသွားမည် ဖြစ်သည်။")
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            edited_df.to_excel(writer, sheet_name='Audited_Ledger', index=False)
            
            # openpyxl ဖြင့် Excel Sheet ထဲဝင်ပြီး အရောင်လိုက်ခြယ်ခြင်း
            workbook = writer.book
            worksheet = writer.sheets['Audited_Ledger']
            
            # အရောင်သတ်မှတ်ချက်များ (Excel Fill)
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Excel Row များကို တစ်ကြောင်းချင်းပတ်ပြီး စစ်ဆေးအရောင်ခြယ်ခြင်း (Row 1 သည် Header မို့ Row 2 ကစသည်)
            for i, (idx, row) in enumerate(edited_df.iterrows(), start=2):
                # 1. Unbalanced -> Red
                if ((row['Debit'] == row['Credit']) & (row['Debit'] != 0)) or ((row['Debit'] == 0) & (row['Credit'] == 0)):
                    for col_num in range(1, len(edited_df.columns) + 1):
                        worksheet.cell(row=i, column=col_num).fill = red_fill
                # 2. Missing Account -> Gray
                elif pd.isna(row['Account_Name']) or row['Account_Name'] in ['nan', '']:
                    for col_num in range(1, len(edited_df.columns) + 1):
                        worksheet.cell(row=i, column=col_num).fill = gray_fill
                # 3. Duplicate Voucher -> Yellow
                elif row['Voucher_No'] in duplicated_vouchers:
                    for col_num in range(1, len(edited_df.columns) + 1):
                        worksheet.cell(row=i, column=col_num).fill = yellow_fill
                        
        st.download_button(
            label="🚀 အရောင်ပါဝင်သော Final Audit Report ကို ဒေါင်းလုဒ်လုပ်ရန်",
            data=buffer.getvalue(),
            file_name="Color_Coded_Audit_Ledger.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        st.error(f"Error တက်သွားပါသည်။ ကော်လံရွေးချယ်မှု ပြန်စစ်ပေးပါ။ အသေးစိတ်: {e}")
else:
    st.info("💡 စတင်ရန် ဘယ်ဘက် Sidebar မှတစ်ဆင့် သင့်ရဲ့ Excel Ledger ဖိုင်ကို တင်သွင်း (Upload) ပေးပါ။")
